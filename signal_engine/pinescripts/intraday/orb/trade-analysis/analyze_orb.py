#!/usr/bin/env python3
"""ORB Intraday Strategy — Signal Performance Analyzer

Parses Telegram channel export and broker trade history (XLSX)
to produce a comprehensive performance report.

Usage:
    python3 analyze_orb.py [telegram_export.json]

    If no file is given, the latest orb-telegram-export-*.json in the
    same directory is used automatically.

    Input naming convention : orb-telegram-export-{YYYY-QX}.json
    Output naming convention : orb-performance-{YYYY-QX}.md

    For a new forward-testing period:
        1. Export the Telegram channel: save as orb-telegram-export-2026-Q2.json
        2. Run: python3 analyze_orb.py orb-telegram-export-2026-Q2.json
        3. Report is written to:  orb-performance-2026-Q2.md
"""

import json
import re
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

# ---------------------------------------------------------------------------
# 1. Message loader
# ---------------------------------------------------------------------------

def load_messages(path: str) -> list[dict]:
    """Load Telegram export and return list of (msg_id, date, text) dicts."""
    with open(path) as f:
        data = json.load(f)

    results = []
    for m in data["messages"]:
        if m["type"] != "message":
            continue
        raw = m.get("text", "")
        if isinstance(raw, list):
            text = "".join(p if isinstance(p, str) else p.get("text", "") for p in raw)
        else:
            text = str(raw)
        results.append({
            "msg_id": m["id"],
            "date": m["date"][:10],       # YYYY-MM-DD
            "datetime": m["date"],
            "text": text,
        })
    return results


# ---------------------------------------------------------------------------
# 2. Entry parser
# ---------------------------------------------------------------------------

# v1: "ORB LONG 🟢\n\nSymbol: FEDERALBNK\nEntry: 283.25\nTarget: 284.69 (T1)\nStop Loss: 281.81\n..."
RE_ENTRY_V1 = re.compile(
    r"ORB\s+(LONG|SHORT)\s.*?"
    r"Symbol:\s*(\w+)\s*\n"
    r"Entry:\s*([\d.]+)\s*\n"
    r"Target:\s*([\d.]+).*?\n"
    r"Stop Loss:\s*([\d.]+)\s*\n"
    r"Risk:\s*[+-]?([\d.]+)\s*\|\s*Reward:\s*[+-]?([\d.]+)\s*\|\s*R:R:\s*[\d.]+\s*\(1:([\d.]+)\)"
    r".*?Time:\s*(\d{2}:\d{2})\s*IST",
    re.DOTALL,
)

# v2: "🟢 ORB LONG | JSWENERGY\n---\nEntry: 495.45\nTarget: 502.52\nSL: 490.74\n---\n..."
RE_ENTRY_V2 = re.compile(
    r"ORB\s+(LONG|SHORT)\s*\|\s*(\w+)\s*\n"
    r"[-]+\s*\n"
    r"Entry:\s*([\d.]+)\s*\n"
    r"Target:\s*([\d.]+)\s*\n"
    r"SL:\s*([\d.]+)\s*\n"
    r"[-]+\s*\n"
    r"Risk:\s*([\d.]+)\s*\|\s*Reward:\s*([\d.]+)\s*\n"
    r"R:R\s*1:([\d.]+)\s*\n"
    r"[-]+\s*\n"
    r"(\d{2}:\d{2})\s*IST",
    re.DOTALL,
)


def parse_entries(messages: list[dict]) -> list[dict]:
    entries = []
    for m in messages:
        t = m["text"]
        if "ORB LONG" not in t and "ORB SHORT" not in t:
            continue

        match = RE_ENTRY_V2.search(t) or RE_ENTRY_V1.search(t)
        if not match:
            print(f"WARN: unparsed entry msg_id={m['msg_id']}: {t[:80]}")
            continue

        direction, symbol, entry, target, sl, risk, reward, rr, time = match.groups()
        entries.append({
            "msg_id": m["msg_id"],
            "date": m["date"],
            "datetime": m["datetime"],
            "symbol": symbol,
            "direction": direction,
            "entry_price": float(entry),
            "target": float(target),
            "sl": float(sl),
            "risk": float(risk),
            "reward": float(reward),
            "rr_ratio": float(rr),
            "signal_time": time,
            "risk_pct": abs(float(entry) - float(sl)) / float(entry) * 100,
        })
    return entries


# ---------------------------------------------------------------------------
# 3. Exit parser
# ---------------------------------------------------------------------------

# TP (v1): "TAKE PROFIT (TP1)\n\nSymbol: ITC\nType: SHORT ▼\n..."
RE_EXIT_TP_V1 = re.compile(
    r"TAKE PROFIT \(TP1\)\s*\n\n"
    r"Symbol:\s*(\w+)\s*\n"
    r"Type:\s*(LONG|SHORT)\s*.*?\n"
    r".*?Entry:\s*([\d.]+)\s*\n"
    r".*?Exit:\s*([\d.]+)\s*\n\n"
    r"Outcome:\s*\n"
    r"P&L:\s*[+-]?[\d.]+\s*\(([+-]?[\d.]+)%\)",
    re.DOTALL,
)

# TP (v2): "✅ TP1 HIT | NATIONALUM\n---\n🟢 LONG | Entry: 381.2\nExit: 385.65\n---\n..."
RE_EXIT_TP_V2 = re.compile(
    r"(TP[\d.]+)\s+HIT\s*\|\s*(\w+)\s*\n"
    r"[-]+\s*\n"
    r".*?(LONG|SHORT)\s*\|\s*Entry:\s*([\d.]+)\s*\n"
    r"Exit:\s*([\d.]+)\s*\n"
    r"[-]+\s*\n"
    r"Profit:\s*[+-]?[\d.]+\s*\(([+-]?[\d.]+)%\)",
    re.DOTALL,
)

# SL: "❌ SL HIT | ITC\n---\n🟢 LONG | Entry: 324.45\nExit: 322.38\n---\nLoss: -2.07 (-0.6%)"
RE_EXIT_SL = re.compile(
    r"SL HIT\s*\|\s*(\w+)\s*\n"
    r"[-]+\s*\n"
    r".*?(LONG|SHORT)\s*\|\s*Entry:\s*([\d.]+)\s*\n"
    r"Exit:\s*([\d.]+)\s*\n"
    r"[-]+\s*\n"
    r"Loss:\s*[+-]?[\d.]+\s*\(([+-]?[\d.]+)%\)",
    re.DOTALL,
)

# TIME EXIT: "⏰ TIME EXIT | FEDERALBNK\n---\n🟢 LONG | Entry: 286.45\nExit: 287.5\n---\nP&L: ..."
RE_EXIT_TIME = re.compile(
    r"TIME EXIT\s*\|\s*(\w+)\s*\n"
    r"[-]+\s*\n"
    r".*?(LONG|SHORT)\s*\|\s*Entry:\s*([\d.]+)\s*\n"
    r"Exit:\s*([\d.]+)\s*\n"
    r"[-]+\s*\n"
    r"P&L:\s*[+-]?[\d.]+\s*\(([+-]?[\d.]+)%\)",
    re.DOTALL,
)


def parse_exits(messages: list[dict]) -> list[dict]:
    exits = []
    for m in messages:
        t = m["text"]
        exit_rec = None

        if "TAKE PROFIT (TP1)" in t:
            match = RE_EXIT_TP_V1.search(t)
            if match:
                symbol, direction, entry, exit_p, pnl = match.groups()
                exit_rec = {"exit_type": "TP1", "symbol": symbol, "direction": direction,
                            "entry_price": float(entry), "exit_price": float(exit_p), "pnl_pct": float(pnl)}

        elif "TP1.5 HIT" in t or "TP1 HIT" in t or "TP2 HIT" in t or "TP3 HIT" in t:
            match = RE_EXIT_TP_V2.search(t)
            if match:
                tp_level, symbol, direction, entry, exit_p, pnl = match.groups()
                exit_rec = {"exit_type": tp_level, "symbol": symbol, "direction": direction,
                            "entry_price": float(entry), "exit_price": float(exit_p), "pnl_pct": float(pnl)}

        elif "SL HIT" in t:
            match = RE_EXIT_SL.search(t)
            if match:
                symbol, direction, entry, exit_p, pnl = match.groups()
                exit_rec = {"exit_type": "SL", "symbol": symbol, "direction": direction,
                            "entry_price": float(entry), "exit_price": float(exit_p), "pnl_pct": float(pnl)}

        elif "TIME EXIT" in t:
            match = RE_EXIT_TIME.search(t)
            if match:
                symbol, direction, entry, exit_p, pnl = match.groups()
                exit_rec = {"exit_type": "TIME_EXIT", "symbol": symbol, "direction": direction,
                            "entry_price": float(entry), "exit_price": float(exit_p), "pnl_pct": float(pnl)}

        if exit_rec:
            exit_rec["msg_id"] = m["msg_id"]
            exit_rec["date"] = m["date"]
            exit_rec["datetime"] = m["datetime"]
            # Extract exit time
            time_match = re.search(r"(\d{2}:\d{2})\s*IST", t)
            exit_rec["exit_time"] = time_match.group(1) if time_match else ""
            exits.append(exit_rec)
        elif "ORB LONG" not in t and "ORB SHORT" not in t:
            print(f"WARN: unparsed exit msg_id={m['msg_id']}: {t[:80]}")

    return exits


# ---------------------------------------------------------------------------
# 4. Trade matcher — link exits to entries, handle multi-TP
# ---------------------------------------------------------------------------

def match_trades(entries: list[dict], exits: list[dict]) -> list[dict]:
    """Match exits to entries. Each entry becomes one trade record.
    Multi-TP exits (TP1, TP1.5, TP2, TP3) for the same entry are linked.
    The 'final_exit' is the highest TP reached; for SL/TE it's just that."""

    # Index entries by (date, symbol, direction, entry_price)
    entry_map = {}
    for e in entries:
        key = (e["date"], e["symbol"], e["direction"], e["entry_price"])
        entry_map[key] = {**e, "exits": [], "final_exit_type": None, "final_pnl_pct": None}

    orphan_exits = []
    for ex in exits:
        key = (ex["date"], ex["symbol"], ex["direction"], ex["entry_price"])
        if key in entry_map:
            entry_map[key]["exits"].append(ex)
        else:
            orphan_exits.append(ex)

    if orphan_exits:
        print(f"WARN: {len(orphan_exits)} orphan exits (no matching entry):")
        for o in orphan_exits:
            print(f"  {o['date']} {o['symbol']} {o['direction']} entry={o['entry_price']} {o['exit_type']}")

    # Determine final exit for each trade
    tp_order = {"TP1": 1, "TP1.5": 2, "TP2": 3, "TP3": 4}
    trades = []
    for key, trade in entry_map.items():
        ex_list = trade["exits"]
        if not ex_list:
            # Entry with no exit — could be end of data
            trade["final_exit_type"] = "NO_EXIT"
            trade["final_pnl_pct"] = 0.0
        elif len(ex_list) == 1:
            trade["final_exit_type"] = ex_list[0]["exit_type"]
            trade["final_pnl_pct"] = ex_list[0]["pnl_pct"]
            trade["exit_price"] = ex_list[0]["exit_price"]
            trade["exit_time"] = ex_list[0]["exit_time"]
        else:
            # Multi-TP: find highest TP level reached
            sl_or_te = [e for e in ex_list if e["exit_type"] in ("SL", "TIME_EXIT")]
            tps = [e for e in ex_list if e["exit_type"].startswith("TP")]
            if tps:
                best_tp = max(tps, key=lambda e: tp_order.get(e["exit_type"], 0))
                trade["final_exit_type"] = best_tp["exit_type"]
                trade["final_pnl_pct"] = best_tp["pnl_pct"]
                trade["exit_price"] = best_tp["exit_price"]
                trade["exit_time"] = best_tp["exit_time"]
            elif sl_or_te:
                trade["final_exit_type"] = sl_or_te[0]["exit_type"]
                trade["final_pnl_pct"] = sl_or_te[0]["pnl_pct"]
                trade["exit_price"] = sl_or_te[0]["exit_price"]
                trade["exit_time"] = sl_or_te[0]["exit_time"]

        # TP progression tracking
        trade["tp_levels_hit"] = sorted(
            [e["exit_type"] for e in ex_list if e["exit_type"].startswith("TP")],
            key=lambda x: tp_order.get(x, 0)
        )

        trades.append(trade)

    return trades


# ---------------------------------------------------------------------------
# 5. XLSX trade history loader
# ---------------------------------------------------------------------------

def load_xlsx_trades(path: str) -> list[dict]:
    """Load broker trade history from XLSX."""
    try:
        import openpyxl
    except ImportError:
        print("WARN: openpyxl not available, skipping XLSX analysis")
        return []

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Find header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        if row[0] and str(row[0]).strip() == "Trade Date":
            header_row = i
            break

    if not header_row:
        print("WARN: Could not find header row in XLSX")
        return []

    trades = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        date_str = str(row[0]).strip() if row[0] else ""
        if not date_str or "-" not in date_str:
            continue
        # Validate it's a date
        try:
            dt = datetime.strptime(date_str, "%d-%m-%Y")
        except ValueError:
            continue

        symbol = str(row[3]).replace("-EQ", "").strip() if row[3] else ""
        trades.append({
            "date": dt.strftime("%Y-%m-%d"),
            "exchange": str(row[1]).strip() if row[1] else "",
            "side": str(row[2]).strip() if row[2] else "",  # Buy/Sell
            "symbol": symbol,
            "qty": int(row[4]) if row[4] else 0,
            "price": float(row[5]) if row[5] else 0.0,
            "trade_id": str(row[6]).strip() if row[6] else "",
        })

    return trades


# ---------------------------------------------------------------------------
# 6. Analysis functions
# ---------------------------------------------------------------------------

def compute_overall_metrics(trades: list[dict]) -> dict:
    """Compute overall performance metrics."""
    completed = [t for t in trades if t["final_exit_type"] != "NO_EXIT"]
    wins = [t for t in completed if t["final_pnl_pct"] > 0]
    losses = [t for t in completed if t["final_pnl_pct"] < 0]
    flat = [t for t in completed if t["final_pnl_pct"] == 0]

    total_pnl = sum(t["final_pnl_pct"] for t in completed)
    avg_pnl = total_pnl / len(completed) if completed else 0

    avg_win = sum(t["final_pnl_pct"] for t in wins) / len(wins) if wins else 0
    avg_loss = sum(t["final_pnl_pct"] for t in losses) / len(losses) if losses else 0
    wl_ratio = abs(avg_win / avg_loss) if avg_loss else 0

    # Win/loss streaks
    max_win_streak = max_loss_streak = cur_win = cur_loss = 0
    for t in sorted(completed, key=lambda x: x["datetime"]):
        if t["final_pnl_pct"] > 0:
            cur_win += 1
            cur_loss = 0
        elif t["final_pnl_pct"] < 0:
            cur_loss += 1
            cur_win = 0
        else:
            cur_win = cur_loss = 0
        max_win_streak = max(max_win_streak, cur_win)
        max_loss_streak = max(max_loss_streak, cur_loss)

    # Daily stats
    daily = defaultdict(lambda: {"trades": 0, "pnl": 0.0})
    for t in completed:
        daily[t["date"]]["trades"] += 1
        daily[t["date"]]["pnl"] += t["final_pnl_pct"]

    winning_days = sum(1 for d in daily.values() if d["pnl"] > 0)
    losing_days = sum(1 for d in daily.values() if d["pnl"] < 0)
    trading_days = len(daily)
    avg_signals_day = len(completed) / trading_days if trading_days else 0

    return {
        "total_entries": len(trades),
        "total_completed": len(completed),
        "wins": len(wins),
        "losses": len(losses),
        "flat": len(flat),
        "win_rate": len(wins) / len(completed) * 100 if completed else 0,
        "total_pnl": total_pnl,
        "avg_pnl": avg_pnl,
        "avg_win": avg_win,
        "avg_loss": avg_loss,
        "wl_ratio": wl_ratio,
        "max_win_streak": max_win_streak,
        "max_loss_streak": max_loss_streak,
        "winning_days": winning_days,
        "losing_days": losing_days,
        "trading_days": trading_days,
        "avg_signals_day": avg_signals_day,
    }


def compute_direction_metrics(trades: list[dict]) -> dict:
    result = {}
    for direction in ("LONG", "SHORT"):
        dt = [t for t in trades if t["direction"] == direction and t["final_exit_type"] != "NO_EXIT"]
        wins = [t for t in dt if t["final_pnl_pct"] > 0]
        losses = [t for t in dt if t["final_pnl_pct"] < 0]
        te = [t for t in dt if t["final_exit_type"] == "TIME_EXIT"]
        total_pnl = sum(t["final_pnl_pct"] for t in dt)
        result[direction] = {
            "trades": len(dt),
            "wins": len(wins),
            "losses": len(losses),
            "te": len(te),
            "wr": len(wins) / len(dt) * 100 if dt else 0,
            "total_pnl": total_pnl,
            "avg_pnl": total_pnl / len(dt) if dt else 0,
        }
    return result


def compute_tp_distribution(trades: list[dict]) -> dict:
    """Count final exit types + TP progression funnel."""
    completed = [t for t in trades if t["final_exit_type"] != "NO_EXIT"]
    counts = defaultdict(int)
    for t in completed:
        counts[t["final_exit_type"]] += 1

    # TP progression: of trades that hit TP1, how many went to TP1.5, etc.
    tp1_hit = [t for t in completed if "TP1" in t["tp_levels_hit"]]
    tp15_hit = [t for t in completed if "TP1.5" in t["tp_levels_hit"]]
    tp2_hit = [t for t in completed if "TP2" in t["tp_levels_hit"]]
    tp3_hit = [t for t in completed if "TP3" in t["tp_levels_hit"]]

    return {
        "counts": dict(counts),
        "total": len(completed),
        "funnel": {
            "tp1": len(tp1_hit),
            "tp1_to_tp15": len(tp15_hit),
            "tp15_to_tp2": len(tp2_hit),
            "tp2_to_tp3": len(tp3_hit),
        }
    }


def compute_day_of_week(trades: list[dict]) -> list[dict]:
    days = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    completed = [t for t in trades if t["final_exit_type"] != "NO_EXIT"]
    result = []
    for i, day_name in enumerate(days):
        dt = [t for t in completed if datetime.strptime(t["date"], "%Y-%m-%d").weekday() == i]
        wins = [t for t in dt if t["final_pnl_pct"] > 0]
        losses = [t for t in dt if t["final_pnl_pct"] < 0]
        pnl = sum(t["final_pnl_pct"] for t in dt)
        result.append({
            "day": day_name,
            "trades": len(dt),
            "wins": len(wins),
            "losses": len(losses),
            "wr": len(wins) / len(dt) * 100 if dt else 0,
            "pnl": pnl,
        })
    return result


def compute_entry_time_dist(trades: list[dict]) -> list[dict]:
    completed = [t for t in trades if t["final_exit_type"] != "NO_EXIT"]
    hours = defaultdict(lambda: {"count": 0, "wins": 0, "losses": 0, "pnl": 0.0})
    for t in completed:
        h = t["signal_time"][:2] + ":xx"
        hours[h]["count"] += 1
        if t["final_pnl_pct"] > 0:
            hours[h]["wins"] += 1
        elif t["final_pnl_pct"] < 0:
            hours[h]["losses"] += 1
        hours[h]["pnl"] += t["final_pnl_pct"]

    result = []
    for h in sorted(hours.keys()):
        d = hours[h]
        d["hour"] = h
        d["wr"] = d["wins"] / d["count"] * 100 if d["count"] else 0
        result.append(d)
    return result


def compute_rr_distribution(trades: list[dict]) -> dict:
    rrs = [t["rr_ratio"] for t in trades]
    if not rrs:
        return {}
    buckets = defaultdict(int)
    for r in rrs:
        if r < 1.0:
            buckets["0.8-1.0"] += 1
        elif r < 1.2:
            buckets["1.0-1.2"] += 1
        elif r < 1.5:
            buckets["1.2-1.5"] += 1
        else:
            buckets["1.5+"] += 1

    rrs_sorted = sorted(rrs)
    return {
        "buckets": dict(buckets),
        "avg": sum(rrs) / len(rrs),
        "median": rrs_sorted[len(rrs_sorted) // 2],
        "min": min(rrs),
        "max": max(rrs),
    }


def compute_risk_per_trade(trades: list[dict]) -> dict:
    risks = [t["risk_pct"] for t in trades]
    if not risks:
        return {}
    risks_sorted = sorted(risks)
    return {
        "min": min(risks),
        "max": max(risks),
        "avg": sum(risks) / len(risks),
        "median": risks_sorted[len(risks_sorted) // 2],
    }


def compute_per_symbol(trades: list[dict]) -> list[dict]:
    completed = [t for t in trades if t["final_exit_type"] != "NO_EXIT"]
    symbols = defaultdict(lambda: {"wins": 0, "losses": 0, "te": 0, "pnl": 0.0, "count": 0})
    for t in completed:
        s = symbols[t["symbol"]]
        s["count"] += 1
        if t["final_pnl_pct"] > 0:
            s["wins"] += 1
        elif t["final_pnl_pct"] < 0:
            s["losses"] += 1
        else:
            s["te"] += 1
        s["pnl"] += t["final_pnl_pct"]

    result = []
    for sym, d in symbols.items():
        wr = d["wins"] / d["count"] * 100 if d["count"] else 0
        avg_pnl = d["pnl"] / d["count"] if d["count"] else 0
        # Grading
        if wr >= 80 and d["pnl"] > 0:
            grade = "A"
        elif wr >= 60 and d["pnl"] > 0:
            grade = "B"
        elif wr >= 40 or (d["pnl"] > 0 and d["pnl"] < 1.0):
            grade = "C"
        else:
            grade = "D"
        result.append({
            "symbol": sym, "wins": d["wins"], "losses": d["losses"], "te": d["te"],
            "count": d["count"], "wr": wr, "pnl": d["pnl"], "avg_pnl": avg_pnl, "grade": grade,
        })

    return sorted(result, key=lambda x: x["pnl"], reverse=True)


def compute_daily_log(trades: list[dict]) -> list[dict]:
    completed = [t for t in trades if t["final_exit_type"] != "NO_EXIT"]
    daily = defaultdict(lambda: {"trades": 0, "wins": 0, "losses": 0, "te": 0, "pnl": 0.0})
    for t in completed:
        d = daily[t["date"]]
        d["trades"] += 1
        if t["final_pnl_pct"] > 0:
            d["wins"] += 1
        elif t["final_pnl_pct"] < 0:
            d["losses"] += 1
        if t["final_exit_type"] == "TIME_EXIT":
            d["te"] += 1
        d["pnl"] += t["final_pnl_pct"]

    result = []
    for dt in sorted(daily.keys()):
        d = daily[dt]
        result.append({
            "date": dt, **d,
            "result": "+" if d["pnl"] > 0 else ("-" if d["pnl"] < 0 else "="),
        })
    return result


def compute_slippage(trades: list[dict], xlsx_trades: list[dict]) -> dict:
    """Cross-reference signals with live fills to measure slippage."""
    if not xlsx_trades:
        return {}

    # Pair XLSX buy/sell into round trips
    buys = defaultdict(list)
    sells = defaultdict(list)
    for t in xlsx_trades:
        key = (t["date"], t["symbol"])
        if t["side"] == "Buy":
            buys[key].append(t)
        else:
            sells[key].append(t)

    # Match signal entries to XLSX fills
    matched = []
    for trade in trades:
        if trade["final_exit_type"] == "NO_EXIT":
            continue
        key = (trade["date"], trade["symbol"])

        # Determine which XLSX side is the entry
        if trade["direction"] == "LONG":
            entry_fills = buys.get(key, [])
            exit_fills = sells.get(key, [])
        else:
            entry_fills = sells.get(key, [])
            exit_fills = buys.get(key, [])

        if entry_fills:
            # Use the fill closest to signal entry price
            best_fill = min(entry_fills, key=lambda f: abs(f["price"] - trade["entry_price"]))
            entry_slip = abs(best_fill["price"] - trade["entry_price"])
            entry_slip_pct = entry_slip / trade["entry_price"] * 100

            exit_slip_pct = 0.0
            if exit_fills and "exit_price" in trade:
                best_exit = min(exit_fills, key=lambda f: abs(f["price"] - trade["exit_price"]))
                exit_slip = abs(best_exit["price"] - trade["exit_price"])
                exit_slip_pct = exit_slip / trade["exit_price"] * 100

            matched.append({
                "date": trade["date"],
                "symbol": trade["symbol"],
                "direction": trade["direction"],
                "signal_entry": trade["entry_price"],
                "fill_entry": best_fill["price"],
                "entry_slippage": entry_slip,
                "entry_slippage_pct": entry_slip_pct,
                "exit_slippage_pct": exit_slip_pct,
                "qty": best_fill["qty"],
            })

    if not matched:
        return {}

    entry_slips = [m["entry_slippage_pct"] for m in matched]
    exit_slips = [m["exit_slippage_pct"] for m in matched]

    return {
        "matched_trades": len(matched),
        "avg_entry_slippage_pct": sum(entry_slips) / len(entry_slips),
        "max_entry_slippage_pct": max(entry_slips),
        "avg_exit_slippage_pct": sum(exit_slips) / len(exit_slips),
        "details": matched,
    }


# ---------------------------------------------------------------------------
# 7. Report generator
# ---------------------------------------------------------------------------

def generate_report(
    trades: list[dict],
    entries: list[dict],
    exits: list[dict],
    xlsx_trades: list[dict],
    q1_cutoff: str = "2026-03-12",
) -> str:
    """Generate comprehensive Markdown report."""

    lines = []
    def w(s=""):
        lines.append(s)

    # Split into periods
    q1_trades = [t for t in trades if t["date"] < q1_cutoff]
    q2_trades = [t for t in trades if t["date"] >= q1_cutoff]

    all_metrics = compute_overall_metrics(trades)
    q1_metrics = compute_overall_metrics(q1_trades)
    q2_metrics = compute_overall_metrics(q2_trades)

    # Date range
    dates = sorted(set(t["date"] for t in trades))
    first_date = dates[0] if dates else "N/A"
    last_date = dates[-1] if dates else "N/A"

    w("# ORB Signal Performance Analysis - Full Period 2026")
    w()
    w(f"**Period:** {first_date} to {last_date} ({all_metrics['trading_days']} trading days)")
    w("**Source:** Telegram channel `intraday-orb` export")
    w("**Strategy:** orb-strategy-luxy-tg (ORB15, 5-min chart)")
    w("**Configuration:**")
    w("- TP levels: TP1, TP1.5, TP2, TP3 (multi-TP added Mar 12)")
    w("- Stop mode: Smart Adaptive / Scaled ATR")
    w("- Entry cutoff: 11:00 AM IST")
    w("- Time exit: 14:30 IST")
    w("- Gap filter: 2.5% max")
    w("- Min entry time: 9:45 AM")
    w("- ORB range filter: 0.4% - 3.5%")
    w("- Commission: 0.06% (mStock)")
    w("- Index filter: Enabled (from Mar 12)")
    w()
    w("---")
    w()

    # Overall Performance
    m = all_metrics
    w("## Overall Performance")
    w()
    w("| Metric | Value |")
    w("|--------|-------|")
    w(f"| Total entries | {m['total_entries']} |")
    w(f"| Total completed trades | {m['total_completed']} |")
    w(f"| Win Rate | **{m['win_rate']:.1f}%** ({m['wins']}/{m['total_completed']}) |")
    w(f"| Cumulative PnL | **{m['total_pnl']:+.2f}%** (sum of individual trade %) |")
    w(f"| Avg PnL/trade | {m['avg_pnl']:+.3f}% |")
    w(f"| Avg win | {m['avg_win']:+.3f}% |")
    w(f"| Avg loss | {m['avg_loss']:+.3f}% |")
    w(f"| Win/Loss ratio | {m['wl_ratio']:.2f} |")
    w(f"| Expectancy/trade | {m['avg_pnl']:+.3f}% |")
    w(f"| Winning days | {m['winning_days']}/{m['trading_days']} ({m['winning_days']/m['trading_days']*100:.1f}%) |")
    w(f"| Losing days | {m['losing_days']}/{m['trading_days']} ({m['losing_days']/m['trading_days']*100:.1f}%) |")
    w(f"| Max win streak | {m['max_win_streak']} |")
    w(f"| Max loss streak | {m['max_loss_streak']} |")
    w(f"| Avg signals/day | {m['avg_signals_day']:.1f} |")
    w()

    # Period Comparison
    w("## Period Comparison: Q1 vs Post-Q1")
    w()
    w("Config changes applied Mar 12: Index direction filter, Volume MA 20->50, Volume 3-bar window, Commission correction 0.03%->0.06%")
    w()
    w("| Metric | Q1 (Jan 28 - Mar 11) | Post-Q1 (Mar 12 - Apr 6) | Delta |")
    w("|--------|---------------------|--------------------------|-------|")

    def delta(a, b, fmt=".1f", suffix="%"):
        d = b - a
        return f"{d:+{fmt}}{suffix}"

    w(f"| Trades | {q1_metrics['total_completed']} | {q2_metrics['total_completed']} | — |")
    w(f"| Win Rate | {q1_metrics['win_rate']:.1f}% | {q2_metrics['win_rate']:.1f}% | {delta(q1_metrics['win_rate'], q2_metrics['win_rate'])} |")
    w(f"| Cumulative PnL | {q1_metrics['total_pnl']:+.2f}% | {q2_metrics['total_pnl']:+.2f}% | {delta(q1_metrics['total_pnl'], q2_metrics['total_pnl'], '.2f')} |")
    w(f"| Avg PnL/trade | {q1_metrics['avg_pnl']:+.3f}% | {q2_metrics['avg_pnl']:+.3f}% | {delta(q1_metrics['avg_pnl'], q2_metrics['avg_pnl'], '.3f')} |")
    w(f"| Avg win | {q1_metrics['avg_win']:+.3f}% | {q2_metrics['avg_win']:+.3f}% | {delta(q1_metrics['avg_win'], q2_metrics['avg_win'], '.3f')} |")
    w(f"| Avg loss | {q1_metrics['avg_loss']:+.3f}% | {q2_metrics['avg_loss']:+.3f}% | {delta(q1_metrics['avg_loss'], q2_metrics['avg_loss'], '.3f')} |")
    w(f"| Win/Loss ratio | {q1_metrics['wl_ratio']:.2f} | {q2_metrics['wl_ratio']:.2f} | {delta(q1_metrics['wl_ratio'], q2_metrics['wl_ratio'], '.2f', '')} |")
    if q1_metrics['trading_days'] > 0 and q2_metrics['trading_days'] > 0:
        q1_wd_pct = q1_metrics['winning_days'] / q1_metrics['trading_days'] * 100
        q2_wd_pct = q2_metrics['winning_days'] / q2_metrics['trading_days'] * 100
        w(f"| Winning days | {q1_metrics['winning_days']}/{q1_metrics['trading_days']} ({q1_wd_pct:.1f}%) | {q2_metrics['winning_days']}/{q2_metrics['trading_days']} ({q2_wd_pct:.1f}%) | {delta(q1_wd_pct, q2_wd_pct)} |")
    w(f"| Max win streak | {q1_metrics['max_win_streak']} | {q2_metrics['max_win_streak']} | — |")
    w(f"| Max loss streak | {q1_metrics['max_loss_streak']} | {q2_metrics['max_loss_streak']} | — |")
    w(f"| Avg signals/day | {q1_metrics['avg_signals_day']:.1f} | {q2_metrics['avg_signals_day']:.1f} | {delta(q1_metrics['avg_signals_day'], q2_metrics['avg_signals_day'], '.1f', '')} |")
    w()

    # Direction Performance
    dir_metrics = compute_direction_metrics(trades)
    w("## Direction Performance")
    w()
    w("| Direction | Trades | Wins | Losses | TE | WR% | Total PnL | Avg PnL |")
    w("|-----------|--------|------|--------|-----|------|-----------|---------|")
    for d in ("LONG", "SHORT"):
        dm = dir_metrics[d]
        w(f"| {d} | {dm['trades']} | {dm['wins']} | {dm['losses']} | {dm['te']} | {dm['wr']:.1f}% | {dm['total_pnl']:+.2f}% | {dm['avg_pnl']:+.3f}% |")
    w()

    # Direction by period
    q1_dir = compute_direction_metrics(q1_trades)
    q2_dir = compute_direction_metrics(q2_trades)
    w("### Direction by Period")
    w()
    w("| Direction | Q1 WR% | Q2 WR% | Q1 PnL | Q2 PnL |")
    w("|-----------|--------|--------|--------|--------|")
    for d in ("LONG", "SHORT"):
        w(f"| {d} | {q1_dir[d]['wr']:.1f}% | {q2_dir[d]['wr']:.1f}% | {q1_dir[d]['total_pnl']:+.2f}% | {q2_dir[d]['total_pnl']:+.2f}% |")
    w()

    # TP Distribution
    tp_dist = compute_tp_distribution(trades)
    w("## TP Hit Distribution")
    w()
    w("| Exit Type | Count | % of Total |")
    w("|-----------|-------|------------|")
    for exit_type in ["TP1", "TP1.5", "TP2", "TP3", "SL", "TIME_EXIT"]:
        count = tp_dist["counts"].get(exit_type, 0)
        pct = count / tp_dist["total"] * 100 if tp_dist["total"] else 0
        w(f"| {exit_type} | {count} | {pct:.1f}% |")
    w()

    # TP Progression Funnel
    funnel = tp_dist["funnel"]
    w("### TP Progression Funnel")
    w()
    w("How many trades progressed through each TP level:")
    w()
    w("| Level | Count | % of Entries | Progression Rate |")
    w("|-------|-------|-------------|-----------------|")
    tp1_n = funnel["tp1"]
    tp15_n = funnel["tp1_to_tp15"]
    tp2_n = funnel["tp15_to_tp2"]
    tp3_n = funnel["tp2_to_tp3"]
    total_entries = tp_dist["total"]
    w(f"| TP1 reached | {tp1_n} | {tp1_n/total_entries*100:.1f}% | — |")
    w(f"| TP1 -> TP1.5 | {tp15_n} | {tp15_n/total_entries*100:.1f}% | {tp15_n/tp1_n*100:.1f}% of TP1 |" if tp1_n else f"| TP1 -> TP1.5 | {tp15_n} | 0% | — |")
    w(f"| TP1.5 -> TP2 | {tp2_n} | {tp2_n/total_entries*100:.1f}% | {tp2_n/tp15_n*100:.1f}% of TP1.5 |" if tp15_n else f"| TP1.5 -> TP2 | {tp2_n} | 0% | — |")
    w(f"| TP2 -> TP3 | {tp3_n} | {tp3_n/total_entries*100:.1f}% | {tp3_n/tp2_n*100:.1f}% of TP2 |" if tp2_n else f"| TP2 -> TP3 | {tp3_n} | 0% | — |")
    w()

    # Entry Time Distribution
    time_dist = compute_entry_time_dist(trades)
    w("## Entry Time Distribution")
    w()
    w("| Hour | Count | % | WR% | PnL% |")
    w("|------|-------|---|-----|------|")
    for td in time_dist:
        pct = td["count"] / all_metrics["total_completed"] * 100 if all_metrics["total_completed"] else 0
        w(f"| {td['hour']} | {td['count']} | {pct:.1f}% | {td['wr']:.1f}% | {td['pnl']:+.2f}% |")
    w()

    # R:R Distribution
    rr_dist = compute_rr_distribution(trades)
    if rr_dist:
        w("## R:R Ratio Distribution")
        w()
        w("| R:R Range | Count |")
        w("|-----------|-------|")
        for bucket in ["0.8-1.0", "1.0-1.2", "1.2-1.5", "1.5+"]:
            w(f"| {bucket} | {rr_dist['buckets'].get(bucket, 0)} |")
        w(f"| Average | 1:{rr_dist['avg']:.2f} |")
        w(f"| Median | 1:{rr_dist['median']:.2f} |")
        w()

    # Risk Per Trade
    risk_dist = compute_risk_per_trade(trades)
    if risk_dist:
        w("## Risk Per Trade")
        w()
        w("| Metric | Value |")
        w("|--------|-------|")
        w(f"| Min risk % | {risk_dist['min']:.2f}% |")
        w(f"| Max risk % | {risk_dist['max']:.2f}% |")
        w(f"| Avg risk % | {risk_dist['avg']:.2f}% |")
        w(f"| Median | {risk_dist['median']:.2f}% |")
        w()

    # Day of Week
    dow = compute_day_of_week(trades)
    w("## Day-of-Week Performance")
    w()
    w("| Day | Trades | Wins | Losses | WR% | PnL% |")
    w("|-----|--------|------|--------|------|-------|")
    for d in dow:
        bold = "**" if d["wr"] < 55 else ""
        w(f"| {bold}{d['day']}{bold} | {bold}{d['trades']}{bold} | {bold}{d['wins']}{bold} | {bold}{d['losses']}{bold} | {bold}{d['wr']:.1f}%{bold} | {bold}{d['pnl']:+.2f}%{bold} |")
    w()

    # Day of week by period
    q1_dow = compute_day_of_week(q1_trades)
    q2_dow = compute_day_of_week(q2_trades)
    w("### Day-of-Week by Period")
    w()
    w("| Day | Q1 WR% | Q2 WR% | Q1 PnL | Q2 PnL |")
    w("|-----|--------|--------|--------|--------|")
    for q1d, q2d in zip(q1_dow, q2_dow):
        w(f"| {q1d['day']} | {q1d['wr']:.1f}% | {q2d['wr']:.1f}% | {q1d['pnl']:+.2f}% | {q2d['pnl']:+.2f}% |")
    w()

    # Daily Performance Log
    daily_log = compute_daily_log(trades)
    w("## Daily Performance Log")
    w()
    w("| Date | # | W | L | TE | PnL% | Result |")
    w("|------|---|---|---|-----|-------|--------|")
    for d in daily_log:
        w(f"| {d['date']} | {d['trades']} | {d['wins']} | {d['losses']} | {d['te']} | {d['pnl']:+.2f}% | {d['result']} |")
    w()

    # Per-Symbol Performance
    sym_perf = compute_per_symbol(trades)
    w("## Per-Symbol Performance")
    w()
    w("| Symbol | W | L | TE | WR% | CumPnL% | AvgPnL% | Grade |")
    w("|--------|---|---|-----|------|---------|---------|-------|")
    for s in sym_perf:
        w(f"| {s['symbol']} | {s['wins']} | {s['losses']} | {s['te']} | {s['wr']:.1f}% | {s['pnl']:+.2f}% | {s['avg_pnl']:+.3f}% | {s['grade']} |")
    w()

    # Symbol grade comparison Q1 vs Q2
    q1_sym = {s["symbol"]: s for s in compute_per_symbol(q1_trades)}
    q2_sym = {s["symbol"]: s for s in compute_per_symbol(q2_trades)}
    all_syms = sorted(set(list(q1_sym.keys()) + list(q2_sym.keys())))

    w("### Symbol Grade Changes (Q1 vs Post-Q1)")
    w()
    w("| Symbol | Q1 Grade | Q2 Grade | Q1 PnL | Q2 PnL | Change |")
    w("|--------|----------|----------|--------|--------|--------|")
    for sym in all_syms:
        q1s = q1_sym.get(sym, {"grade": "—", "pnl": 0})
        q2s = q2_sym.get(sym, {"grade": "—", "pnl": 0})
        if q1s["grade"] != q2s["grade"] or sym in q1_sym and sym in q2_sym:
            change = ""
            if q1s["grade"] != "—" and q2s["grade"] != "—":
                grade_val = {"A": 4, "B": 3, "C": 2, "D": 1, "—": 0}
                diff = grade_val.get(q2s["grade"], 0) - grade_val.get(q1s["grade"], 0)
                if diff > 0:
                    change = "UP"
                elif diff < 0:
                    change = "DOWN"
                else:
                    change = "="
            elif q1s["grade"] == "—":
                change = "NEW"
            else:
                change = "GONE"

            q1_pnl = f"{q1s['pnl']:+.2f}%" if isinstance(q1s['pnl'], float) else "—"
            q2_pnl = f"{q2s['pnl']:+.2f}%" if isinstance(q2s['pnl'], float) else "—"
            w(f"| {sym} | {q1s['grade']} | {q2s['grade']} | {q1_pnl} | {q2_pnl} | {change} |")
    w()

    w("### Grading Key")
    w("- **A**: WR >= 80% AND positive PnL (core watchlist)")
    w("- **B**: WR >= 60% AND positive PnL (keep, monitor)")
    w("- **C**: WR >= 40% OR marginal PnL (watch closely, may remove)")
    w("- **D**: WR < 40% OR significant negative PnL (remove candidate)")
    w()

    # Execution Quality (Slippage)
    slippage = compute_slippage(trades, xlsx_trades)
    if slippage:
        w("## Execution Quality (Live Trade Cross-Reference)")
        w()
        w(f"**Period:** Mar 12 - Apr 6, 2026 (live broker data)")
        w(f"**Matched trades:** {slippage['matched_trades']}")
        w()
        w("| Metric | Value |")
        w("|--------|-------|")
        w(f"| Avg entry slippage | {slippage['avg_entry_slippage_pct']:.3f}% |")
        w(f"| Max entry slippage | {slippage['max_entry_slippage_pct']:.3f}% |")
        w(f"| Avg exit slippage | {slippage['avg_exit_slippage_pct']:.3f}% |")
        w()

        # Per-trade slippage detail
        w("### Per-Trade Slippage Detail")
        w()
        w("| Date | Symbol | Dir | Signal Entry | Fill Entry | Slip% | Qty |")
        w("|------|--------|-----|-------------|------------|-------|-----|")
        for d in sorted(slippage["details"], key=lambda x: x["date"]):
            w(f"| {d['date']} | {d['symbol']} | {d['direction']} | {d['signal_entry']:.2f} | {d['fill_entry']:.2f} | {d['entry_slippage_pct']:.3f}% | {d['qty']} |")
        w()

    # TP Strategy Analysis (updated with multi-TP data)
    w("## TP Strategy Analysis")
    w()
    w("### Multi-TP Impact Assessment")
    w()

    # Calculate what TP1-only vs multi-TP would return
    completed = [t for t in trades if t["final_exit_type"] != "NO_EXIT"]
    tp1_only_pnl = 0.0
    multi_tp_pnl = 0.0
    for t in completed:
        if t["final_pnl_pct"] > 0:
            # For TP1-only: cap at 1R (use risk_pct as 1R proxy)
            tp1_only_pnl += min(t["risk_pct"], abs(t["final_pnl_pct"]))
            multi_tp_pnl += t["final_pnl_pct"]
        else:
            tp1_only_pnl += t["final_pnl_pct"]
            multi_tp_pnl += t["final_pnl_pct"]

    w(f"| Strategy | Gross PnL% | Per Trade |")
    w(f"|----------|-----------|-----------|")
    w(f"| TP1 only (cap at 1R) | {tp1_only_pnl:+.2f}% | {tp1_only_pnl/len(completed):+.3f}% |")
    w(f"| Multi-TP (actual) | {multi_tp_pnl:+.2f}% | {multi_tp_pnl/len(completed):+.3f}% |")
    w(f"| Multi-TP uplift | {multi_tp_pnl - tp1_only_pnl:+.2f}% | — |")
    w()

    # Strengths & Weaknesses
    w("## Strengths & Weaknesses")
    w()
    w("### Strengths")
    w()

    strengths = []
    if m["win_rate"] > 65:
        strengths.append(f"- High win rate ({m['win_rate']:.1f}%) provides consistent daily returns")
    if m["winning_days"] / m["trading_days"] * 100 > 75:
        strengths.append(f"- {m['winning_days']}/{m['trading_days']} winning days ({m['winning_days']/m['trading_days']*100:.1f}%) — strong day-level consistency")

    # Check short vs long
    if dir_metrics["SHORT"]["wr"] > dir_metrics["LONG"]["wr"] + 5:
        strengths.append(f"- SHORT trades outperform LONG ({dir_metrics['SHORT']['wr']:.1f}% vs {dir_metrics['LONG']['wr']:.1f}% WR)")
    if dir_metrics["LONG"]["wr"] > dir_metrics["SHORT"]["wr"] + 5:
        strengths.append(f"- LONG trades outperform SHORT ({dir_metrics['LONG']['wr']:.1f}% vs {dir_metrics['SHORT']['wr']:.1f}% WR)")

    # Best symbols
    a_grade = [s for s in sym_perf if s["grade"] == "A"]
    if a_grade:
        strengths.append(f"- {len(a_grade)} A-grade symbols form reliable core watchlist")

    # Q1 vs Q2 improvement
    if q2_metrics["win_rate"] > q1_metrics["win_rate"]:
        strengths.append(f"- Post-config-change WR improved: {q1_metrics['win_rate']:.1f}% -> {q2_metrics['win_rate']:.1f}%")

    for s in strengths:
        w(s)
    w()

    w("### Weaknesses")
    w()

    weaknesses = []
    # Friday check
    fri = [d for d in dow if d["day"] == "Fri"][0] if dow else None
    if fri and fri["wr"] < 55:
        weaknesses.append(f"- Friday underperforms: {fri['wr']:.1f}% WR, {fri['pnl']:+.2f}% PnL")

    # D-grade symbols
    d_grade = [s for s in sym_perf if s["grade"] == "D"]
    if d_grade:
        d_loss = sum(s["pnl"] for s in d_grade)
        weaknesses.append(f"- {len(d_grade)} D-grade symbols dragging PnL ({d_loss:+.2f}%): {', '.join(s['symbol'] for s in d_grade)}")

    # Long vs Short gap
    if abs(dir_metrics["LONG"]["wr"] - dir_metrics["SHORT"]["wr"]) > 10:
        weaker = "LONG" if dir_metrics["LONG"]["wr"] < dir_metrics["SHORT"]["wr"] else "SHORT"
        weaknesses.append(f"- {weaker} direction significantly weaker ({dir_metrics[weaker]['wr']:.1f}% WR)")

    # Late entries
    late_entries = [td for td in time_dist if td["hour"] >= "10:xx" and td["wr"] < 60]
    if late_entries:
        for le in late_entries:
            weaknesses.append(f"- {le['hour']} entries have lower WR ({le['wr']:.1f}%)")

    for w_ in weaknesses:
        w(w_)
    w()

    # Recommendations
    w("## Recommendations")
    w()

    recs = []

    # D-grade symbol removal
    if d_grade:
        d_syms = ", ".join(s["symbol"] for s in d_grade)
        d_loss = sum(s["pnl"] for s in d_grade)
        recs.append(f"1. **Remove D-grade symbols** ({d_syms}): saves {abs(d_loss):.2f}% — highest ROI improvement")

    # Friday handling
    if fri and fri["wr"] < 55:
        recs.append(f"2. **Reduce Friday exposure**: Consider halving max_open_positions on Fridays or tightening entry filters")

    # Direction bias
    if dir_metrics["LONG"]["wr"] < 55:
        recs.append(f"3. **Tighten LONG entry filters**: LONG WR ({dir_metrics['LONG']['wr']:.1f}%) is below 55% — consider requiring stronger index confirmation for LONG entries")
    if dir_metrics["SHORT"]["wr"] < 55:
        recs.append(f"3. **Tighten SHORT entry filters**: SHORT WR ({dir_metrics['SHORT']['wr']:.1f}%) is below 55%")

    # Multi-TP assessment
    if tp1_only_pnl < multi_tp_pnl:
        recs.append(f"4. **Multi-TP is working**: +{multi_tp_pnl - tp1_only_pnl:.2f}% uplift over TP1-only. Keep multi-TP execution if operationally feasible")
    else:
        recs.append(f"4. **Stick with TP1-only**: Multi-TP shows no uplift ({multi_tp_pnl - tp1_only_pnl:+.2f}%). Simpler execution = fewer failure modes")

    # C-grade watchlist
    c_grade = [s for s in sym_perf if s["grade"] == "C"]
    if c_grade:
        c_syms = ", ".join(s["symbol"] for s in c_grade)
        recs.append(f"5. **Monitor C-grade symbols** ({c_syms}): borderline performance — review after 20 more trades each")

    for r in recs:
        w(r)
    w()

    # Baseline update
    w("## Baseline for Future Comparison")
    w()
    w("Use these metrics as baseline when evaluating improvements:")
    w()
    w(f"- **Win Rate baseline**: {m['win_rate']:.1f}%")
    w(f"- **Expectancy baseline**: {m['avg_pnl']:+.3f}%/trade")
    w(f"- **Winning days baseline**: {m['winning_days']/m['trading_days']*100:.1f}%")
    w(f"- **Long WR baseline**: {dir_metrics['LONG']['wr']:.1f}%")
    w(f"- **Short WR baseline**: {dir_metrics['SHORT']['wr']:.1f}%")
    if fri:
        w(f"- **Friday WR baseline**: {fri['wr']:.1f}%")
    w()

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _derive_output_name(export_path: Path, trades: list) -> str:
    """Derive report filename from the export filename or data date range."""
    # If export follows orb-telegram-export-{period}.json, mirror it
    stem = export_path.stem  # e.g. "orb-telegram-export-2026-Q1"
    if stem.startswith("orb-telegram-export-"):
        period = stem[len("orb-telegram-export-"):]
        return f"orb-performance-{period}.md"

    # Fallback: derive period from first/last trade date
    if trades:
        dates = sorted(set(t["date"] for t in trades))
        return f"orb-performance-{dates[0]}-to-{dates[-1]}.md"

    return "orb-performance.md"


def main():
    base = Path(__file__).parent

    # Resolve input export file
    if len(sys.argv) > 1:
        export_path = Path(sys.argv[1])
        if not export_path.is_absolute():
            export_path = base / export_path
    else:
        # Auto-detect latest orb-telegram-export-*.json
        candidates = sorted(base.glob("orb-telegram-export-*.json"))
        if not candidates:
            print("ERROR: no orb-telegram-export-*.json found. Pass the export file as an argument.")
            sys.exit(1)
        export_path = candidates[-1]
        print(f"Auto-selected export: {export_path.name}")

    print("Loading Telegram messages...")
    messages = load_messages(str(export_path))
    print(f"  {len(messages)} messages loaded")

    print("Parsing entries...")
    entries = parse_entries(messages)
    print(f"  {len(entries)} entries parsed")

    print("Parsing exits...")
    exits = parse_exits(messages)
    print(f"  {len(exits)} exits parsed")

    print("Matching trades...")
    trades = match_trades(entries, exits)
    completed = [t for t in trades if t["final_exit_type"] != "NO_EXIT"]
    no_exit = [t for t in trades if t["final_exit_type"] == "NO_EXIT"]
    print(f"  {len(completed)} completed trades, {len(no_exit)} no-exit")

    xlsx_path = base / "TradeHistory_06Mar26_to_07Apr26_MA6718246.xlsx"
    xlsx_trades = []
    if xlsx_path.exists():
        print("Loading XLSX trade history...")
        xlsx_trades = load_xlsx_trades(str(xlsx_path))
        print(f"  {len(xlsx_trades)} broker trades loaded")

    print("Generating report...")
    report = generate_report(trades, entries, exits, xlsx_trades)

    out_name = _derive_output_name(export_path, trades)
    out_path = base / out_name
    with open(out_path, "w") as f:
        f.write(report)
    print(f"Report written to {out_path}")

    # Also print summary
    m = compute_overall_metrics(trades)
    print(f"\n=== SUMMARY ===")
    print(f"Period: {sorted(set(t['date'] for t in trades))[0]} to {sorted(set(t['date'] for t in trades))[-1]}")
    print(f"Trades: {m['total_completed']} | WR: {m['win_rate']:.1f}% | PnL: {m['total_pnl']:+.2f}%")
    print(f"Avg PnL: {m['avg_pnl']:+.3f}% | W/L: {m['wl_ratio']:.2f}")


if __name__ == "__main__":
    main()
